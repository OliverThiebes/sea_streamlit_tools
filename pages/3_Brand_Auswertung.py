from __future__ import annotations

import csv
import io
import re
from collections import Counter, defaultdict
from datetime import datetime

import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

# feste Wechselkurse (1 Einheit Waehrung -> EUR)
FX_TO_EUR = {
    "EUR": 1.0,
    "CHF": 1.0731,
    "PLN": 0.2363,
    "CZK": 0.0413,
    "GBP": 1.1372,
    "DKK": 0.134,
    "SEK": 0.089,
}

DEFAULT_EXPORT_BASE_NAME = "Brand_Auswertung"


def _decode_text(raw: bytes, file_name: str) -> str:
    parse_errors: list[str] = []
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError as exc:
            parse_errors.append(f"{encoding}: {exc}")
    raise ValueError(
        f"Datei '{file_name}' konnte nicht gelesen werden. Fehler: {' | '.join(parse_errors)}"
    )


def _sniff_delimiter(sample: str, fallback: str, delimiters: list[str]) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=delimiters)
        return dialect.delimiter
    except Exception:  # noqa: BLE001
        return fallback


def _sanitize_export_base_name(file_name: str) -> str:
    clean = re.sub(r"\.xlsx$", "", (file_name or "").strip(), flags=re.IGNORECASE)
    clean = re.sub(r'[\\/:*?"<>|]', "_", clean)
    clean = re.sub(r"\s+", "_", clean)
    return clean.strip("._")


def _resolve_output_file_name(file_name: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = _sanitize_export_base_name(file_name)
    if not base_name:
        base_name = f"{DEFAULT_EXPORT_BASE_NAME}_{timestamp}"
    return f"{base_name}.xlsx"


def parse_eu_number(value: str | None) -> float:
    if value is None:
        return 0.0
    text = value.strip().replace('"', "")
    if not text or text == "--":
        return 0.0
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_int(value: str | None) -> int:
    return int(parse_eu_number(value))


def get_fx_factor(waehrung: str) -> float:
    code = (waehrung or "EUR").strip().upper()
    factor = FX_TO_EUR.get(code)
    if factor is None:
        return 1.0
    return factor


def lese_brand_counts(feed_raw: bytes, file_name: str) -> Counter[str]:
    text = _decode_text(feed_raw, file_name)
    sample = "\n".join(text.splitlines()[:20])
    delimiter = _sniff_delimiter(sample, "\t", ["\t", ";", ","])
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    if not reader.fieldnames:
        raise ValueError(f"Feed-Datei '{file_name}' ist leer oder hat keinen Header.")

    field_map = {field.strip().lower(): field for field in reader.fieldnames if field}
    brand_column = field_map.get("brand")
    if not brand_column:
        raise ValueError(
            f"Spalte 'brand' wurde im Feed '{file_name}' nicht gefunden. "
            f"Gefundene Spalten: {reader.fieldnames}"
        )

    brand_counts: Counter[str] = Counter()
    for row in reader:
        brand = (row.get(brand_column) or "").strip()
        if brand:
            brand_counts[brand] += 1

    if not brand_counts:
        raise ValueError(f"Feed-Datei '{file_name}' enthaelt keine verwertbaren Brand-Werte.")
    return brand_counts


def get_timeframe_and_linecount(search_text: str) -> tuple[str, str, int]:
    lines = search_text.splitlines()
    title = lines[0].strip() if lines else ""
    timeframe = lines[1].strip() if len(lines) > 1 else ""
    data_lines = max(len(lines) - 3, 0)
    return title, timeframe, data_lines


def baue_brand_token_index(brands: list[str]) -> dict[str, set[str]]:
    index: dict[str, set[str]] = defaultdict(set)
    for brand in brands:
        tokens = re.findall(r"\w+", brand.lower())
        for token in tokens:
            index[token].add(brand)
    return index


def aggregiere_suchbegriffe(
    such_raw: bytes,
    file_name: str,
    brand_counts: Counter[str],
    progress_cb=None,
) -> tuple[dict[str, dict[str, float]], str]:
    text = _decode_text(such_raw, file_name)
    lines = text.splitlines()
    if len(lines) < 3:
        raise ValueError(
            f"Suchbegriffe-Datei '{file_name}' ist ungueltig. "
            "Es werden mindestens 2 Infozeilen plus Header erwartet."
        )

    _, timeframe, total_lines = get_timeframe_and_linecount(text)
    delimiter = _sniff_delimiter(lines[2], ",", [",", ";", "\t"])
    data_reader = csv.DictReader(io.StringIO("\n".join(lines[2:])), delimiter=delimiter)

    if not data_reader.fieldnames:
        raise ValueError(f"Suchbegriffe-Datei '{file_name}' enthaelt keinen Header.")

    required_cols = [
        "Suchbegriff",
        "Interaktionen",
        "Kosten (umgerechnete Währung)",
        "Conversions",
        "Conv.-Wert",
        "Währungscode",
    ]
    missing = [col for col in required_cols if col not in data_reader.fieldnames]
    if missing:
        raise ValueError(
            f"Suchbegriffe-Datei '{file_name}': fehlende Spalten: {', '.join(missing)}"
        )

    stats = {
        brand: {
            "produkte": count,
            "klicks": 0,
            "kosten_eur": 0.0,
            "conversions": 0.0,
            "conv_wert_eur": 0.0,
        }
        for brand, count in brand_counts.items()
    }

    brand_token_index = baue_brand_token_index(list(brand_counts.keys()))
    brands_lower = {brand: brand.lower() for brand in brand_counts.keys()}

    for i, row in enumerate(data_reader, start=1):
        suchbegriff_raw = row["Suchbegriff"] or ""
        suchbegriff = suchbegriff_raw.lower()
        if not suchbegriff:
            continue

        klicks = parse_int(row["Interaktionen"])
        kosten_eur = parse_eu_number(row["Kosten (umgerechnete Währung)"])
        conversions = parse_eu_number(row["Conversions"])
        conv_wert = parse_eu_number(row["Conv.-Wert"])
        waehrung = row.get("Währungscode", "EUR")
        conv_wert_eur = conv_wert * get_fx_factor(waehrung)

        if progress_cb and (i % 1000 == 0 or i == total_lines):
            progress_cb(i, total_lines)

        tokens = re.findall(r"\w+", suchbegriff)
        candidate_brands: set[str] = set()
        for token in tokens:
            if token in brand_token_index:
                candidate_brands.update(brand_token_index[token])

        if not candidate_brands:
            continue

        for brand in candidate_brands:
            if brands_lower[brand] in suchbegriff:
                values = stats[brand]
                values["klicks"] += klicks
                values["kosten_eur"] += kosten_eur
                values["conversions"] += conversions
                values["conv_wert_eur"] += conv_wert_eur

    return stats, timeframe


def lese_ausschlussliste(raw: bytes, file_name: str) -> set[str]:
    text = _decode_text(raw, file_name)
    lines = text.splitlines()
    if len(lines) < 3:
        raise ValueError(
            f"Ausschlussliste '{file_name}' ist ungueltig. "
            "Es werden mindestens 2 Infozeilen plus Header erwartet."
        )

    delimiter = _sniff_delimiter(lines[2], ",", [",", ";", "\t"])
    reader = csv.DictReader(io.StringIO("\n".join(lines[2:])), delimiter=delimiter)

    if not reader.fieldnames:
        raise ValueError(f"Ausschlussliste '{file_name}' enthaelt keinen Header.")
    if "keyword_text" not in reader.fieldnames:
        raise ValueError(
            f"Spalte 'keyword_text' fehlt in der Ausschlussliste '{file_name}'. "
            f"Gefundene Spalten: {reader.fieldnames}"
        )

    excluded: set[str] = set()
    for row in reader:
        keyword = (row.get("keyword_text") or "").strip().lower()
        if keyword:
            excluded.add(keyword)
    return excluded


def berechne_gesamtsummen_liste(rows) -> tuple[int, float, float, float, float, float]:
    total_klicks = 0
    total_kosten = 0.0
    total_convs = 0.0
    total_conv_wert = 0.0
    for _, _, klicks, _, kosten, convs, conv_wert, _ in rows:
        total_klicks += klicks
        total_kosten += kosten
        total_convs += convs
        total_conv_wert += conv_wert

    cpc = total_kosten / total_klicks if total_klicks > 0 else 0.0
    roas = total_conv_wert / total_kosten if total_kosten > 0 else 0.0
    return total_klicks, total_kosten, total_convs, total_conv_wert, cpc, roas


def _set_common_column_widths(ws, widths: list[int]) -> None:
    for col_letter, width in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I"], widths):
        ws.column_dimensions[col_letter].width = width


def _write_result_headers(ws, row: int) -> None:
    headers = [
        "Marke",
        "Produkte",
        "Klicks",
        "CPC",
        "Kosten in EUR",
        "Conversions",
        "Con.-Wert in EUR",
        "ROAS",
        "ROAS Spiegel",
    ]
    for col_idx, head in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=head)
        cell.font = Font(bold=True)


def _write_result_row(ws, row: int, values) -> None:
    brand, produkte, klicks, cpc, kosten, convs, conv_wert, roas = values
    ws[f"A{row}"] = brand
    ws[f"B{row}"] = produkte
    ws[f"C{row}"] = klicks
    ws[f"D{row}"] = cpc
    ws[f"E{row}"] = kosten
    ws[f"F{row}"] = convs
    ws[f"G{row}"] = conv_wert
    ws[f"H{row}"] = roas
    ws[f"I{row}"] = roas

    ws[f"D{row}"].number_format = '#,##0.00 "EUR"'
    ws[f"E{row}"].number_format = '#,##0.00 "EUR"'
    ws[f"F{row}"].number_format = "#,##0.00"
    ws[f"G{row}"].number_format = '#,##0.00 "EUR"'
    ws[f"H{row}"].number_format = "0%"
    ws[f"I{row}"].number_format = "0%"


def erstelle_sheet_shopmarken_gesamt(wb: Workbook, ergebnisse, timeframe: str) -> None:
    ws = wb.active
    ws.title = "Shopmarken Gesamt"
    _set_common_column_widths(ws, [25, 10, 10, 10, 15, 15, 18, 10, 15])

    ws["A1"] = "Google | Kosten Markensuchanfragen"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:I1")

    ws["A2"] = timeframe or "Zeitraum nicht gefunden"
    ws["A2"].font = Font(size=11, italic=True)
    ws.merge_cells("A2:I2")

    ws["A4"] = (
        "Auswertung der Suchanfragen ueber alle Laender hinweg, die Marken aus dem Shop "
        "enthalten (Markenliste basierend auf DE)."
    )
    ws["A4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A4:H4")

    total_klicks, total_kosten, total_convs, total_conv_wert, total_cpc, total_roas = (
        berechne_gesamtsummen_liste(ergebnisse)
    )

    row = 6
    _write_result_row(
        ws,
        row,
        ("GESAMT", "", total_klicks, total_cpc, total_kosten, total_convs, total_conv_wert, total_roas),
    )
    ws[f"A{row}"].font = Font(bold=True)

    row = 10
    _write_result_headers(ws, row)
    for result in ergebnisse:
        row += 1
        _write_result_row(ws, row, result)

    if row >= 11:
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.conditional_formatting.add(
            f"I11:I{row}",
            CellIsRule(operator="lessThan", formula=["2"], fill=red_fill),
        )


def erstelle_sheet_low_performer(wb: Workbook, ergebnisse, timeframe: str) -> None:
    ws = wb.create_sheet(title="Low Performer")
    _set_common_column_widths(ws, [25, 10, 10, 10, 15, 15, 18, 10, 15])

    ws["A1"] = "Google | Kosten Markensuchanfragen | Low Performer mit einem ROAS < 200%"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:I1")

    ws["A2"] = timeframe or "Zeitraum nicht gefunden"
    ws["A2"].font = Font(size=11, italic=True)
    ws.merge_cells("A2:I2")

    ws["A4"] = "Kostenblock der Marken mit einem ROAS unter 200%"
    ws.merge_cells("A4:H4")

    low_performers = [row for row in ergebnisse if row[7] < 2.0]
    total_klicks, total_kosten, total_convs, total_conv_wert, total_cpc, total_roas = (
        berechne_gesamtsummen_liste(low_performers)
    )

    row = 6
    _write_result_row(
        ws,
        row,
        ("GESAMT", "", total_klicks, total_cpc, total_kosten, total_convs, total_conv_wert, total_roas),
    )
    ws[f"A{row}"].font = Font(bold=True)

    row = 10
    _write_result_headers(ws, row)
    for result in low_performers:
        row += 1
        _write_result_row(ws, row, result)

    if row >= 11:
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.conditional_formatting.add(
            f"I11:I{row}",
            CellIsRule(operator="lessThan", formula=["2"], fill=red_fill),
        )


def erstelle_sheet_top_performer(wb: Workbook, ergebnisse, timeframe: str) -> None:
    ws = wb.create_sheet(title="Top Performer")
    _set_common_column_widths(ws, [25, 10, 10, 10, 15, 15, 18, 10, 15])

    ws["A1"] = "Google | Kosten Markensuchanfragen | Top Performer mit einem ROAS > 200%"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:I1")

    ws["A2"] = timeframe or "Zeitraum nicht gefunden"
    ws["A2"].font = Font(size=11, italic=True)
    ws.merge_cells("A2:I2")

    ws["A4"] = "Marken mit einem ROAS ueber 200 Prozent, sortiert nach Conversionwert in EUR."
    ws.merge_cells("A4:H4")

    top_performers = [row for row in ergebnisse if row[7] > 2.0]
    top_performers.sort(key=lambda value: value[6], reverse=True)
    total_klicks, total_kosten, total_convs, total_conv_wert, total_cpc, total_roas = (
        berechne_gesamtsummen_liste(top_performers)
    )

    row = 6
    _write_result_row(
        ws,
        row,
        ("GESAMT", "", total_klicks, total_cpc, total_kosten, total_convs, total_conv_wert, total_roas),
    )
    ws[f"A{row}"].font = Font(bold=True)

    row = 10
    _write_result_headers(ws, row)
    for result in top_performers:
        row += 1
        _write_result_row(ws, row, result)

    if row >= 11:
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.conditional_formatting.add(
            f"I11:I{row}",
            CellIsRule(operator="greaterThan", formula=["6"], fill=green_fill),
        )


def erstelle_sheet_weniger_als_10(wb: Workbook, stats, brand_counts, exclude_set: set[str]) -> None:
    ws = wb.create_sheet(title="weniger als 10 Produkte")
    _set_common_column_widths(ws, [25, 10, 10, 10, 15, 15, 18, 10, 25])

    ws["A1"] = "Marken mit weniger als 10 Produkten im Feed"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:I1")

    row = 3
    headers = [
        "Marke",
        "Produkte",
        "Klicks",
        "CPC",
        "Kosten in EUR",
        "Conversions",
        "Con.-Wert in EUR",
        "ROAS",
        "Ueber Liste ausgeschlossen",
    ]
    for col_idx, head in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=head)
        cell.font = Font(bold=True)

    for brand, data in stats.items():
        produkte = brand_counts.get(brand, data["produkte"])
        if produkte >= 10:
            continue

        klicks = data["klicks"]
        kosten = data["kosten_eur"]
        convs = data["conversions"]
        conv_wert = data["conv_wert_eur"]
        cpc = kosten / klicks if klicks > 0 else 0.0
        roas = conv_wert / kosten if kosten > 0 else 0.0

        row += 1
        ws[f"A{row}"] = brand
        ws[f"B{row}"] = produkte
        ws[f"C{row}"] = klicks
        ws[f"D{row}"] = cpc
        ws[f"E{row}"] = kosten
        ws[f"F{row}"] = convs
        ws[f"G{row}"] = conv_wert
        ws[f"H{row}"] = roas
        ws[f"I{row}"] = "Ueber Liste ausgeschlossen" if brand.lower() in exclude_set else ""

        ws[f"D{row}"].number_format = '#,##0.00 "EUR"'
        ws[f"E{row}"].number_format = '#,##0.00 "EUR"'
        ws[f"F{row}"].number_format = "#,##0.00"
        ws[f"G{row}"].number_format = '#,##0.00 "EUR"'
        ws[f"H{row}"].number_format = "0%"


def erstelle_sheet_ausgeschlossen_ueber_9(
    wb: Workbook,
    stats,
    brand_counts,
    exclude_set: set[str],
) -> None:
    ws = wb.create_sheet(title="Ausgeschlossen aber >9 Produkte")

    for col_letter, width in zip(["A", "B", "C", "D", "E", "F", "G", "H"], [25, 10, 10, 10, 15, 15, 18, 10]):
        ws.column_dimensions[col_letter].width = width

    ws["A1"] = "Ausgeschlossene Marken mit mindestens 10 Produkten im Feed"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:H1")

    row = 3
    headers = [
        "Marke",
        "Produkte",
        "Klicks",
        "CPC",
        "Kosten in EUR",
        "Conversions",
        "Con.-Wert in EUR",
        "ROAS",
    ]
    for col_idx, head in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=head)
        cell.font = Font(bold=True)

    for brand, data in stats.items():
        produkte = brand_counts.get(brand, data["produkte"])
        if produkte <= 9 or brand.lower() not in exclude_set:
            continue

        klicks = data["klicks"]
        kosten = data["kosten_eur"]
        convs = data["conversions"]
        conv_wert = data["conv_wert_eur"]
        cpc = kosten / klicks if klicks > 0 else 0.0
        roas = conv_wert / kosten if kosten > 0 else 0.0

        row += 1
        ws[f"A{row}"] = brand
        ws[f"B{row}"] = produkte
        ws[f"C{row}"] = klicks
        ws[f"D{row}"] = cpc
        ws[f"E{row}"] = kosten
        ws[f"F{row}"] = convs
        ws[f"G{row}"] = conv_wert
        ws[f"H{row}"] = roas

        ws[f"D{row}"].number_format = '#,##0.00 "EUR"'
        ws[f"E{row}"].number_format = '#,##0.00 "EUR"'
        ws[f"F{row}"].number_format = "#,##0.00"
        ws[f"G{row}"].number_format = '#,##0.00 "EUR"'
        ws[f"H{row}"].number_format = "0%"


def schreibe_excel_bytes(ergebnisse, stats, brand_counts, exclude_set, timeframe: str) -> bytes:
    wb = Workbook()
    erstelle_sheet_shopmarken_gesamt(wb, ergebnisse, timeframe)
    erstelle_sheet_low_performer(wb, ergebnisse, timeframe)
    erstelle_sheet_top_performer(wb, ergebnisse, timeframe)
    erstelle_sheet_weniger_als_10(wb, stats, brand_counts, exclude_set)
    erstelle_sheet_ausgeschlossen_ueber_9(wb, stats, brand_counts, exclude_set)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _render_ui() -> None:
    st.set_page_config(page_title="Brand Auswertung", layout="centered")
    st.title("Brand Auswertung")
    st.markdown(
        """
Diese Analyse prueft, wie stark Marken aus dem Feed in den Suchbegriffen auftreten und
erstellt daraus eine Excel-Auswertung mit mehreren Tabs:

- `Shopmarken Gesamt`: Gesamtauswertung aller erkannten Marken.
- `Low Performer`: Marken mit ROAS unter 200%.
- `Top Performer`: Marken mit ROAS ueber 200%, sortiert nach Conversionwert.
- `weniger als 10 Produkte`: Marken mit kleinem Sortiment inkl. Ausschluss-Check.
- `Ausgeschlossen aber >9 Produkte`: ausgeschlossene Marken mit mindestens 10 Produkten.

**Benötigte Uploads**
- `Feed`: Produktfeed mit Spalte `brand`.
- `Suchbegriffe`: Google Ads Suchbegriffe-Bericht mit den Spalten
  `Suchbegriff`, `Interaktionen`, `Kosten (umgerechnete Währung)`, `Conversions`,
  `Conv.-Wert`, `Währungscode`.
- `Ausschlussliste`: Bericht mit Spalte `keyword_text`.
        """.strip()
    )

    output_file_name = st.text_input(
        "Dateiname fuer den Export (ohne .xlsx)",
        value=DEFAULT_EXPORT_BASE_NAME,
        key="brand_export_name",
    )

    feed_upload = st.file_uploader(
        "Feed-Datei",
        type=["csv", "txt", "tsv"],
        accept_multiple_files=False,
        key="brand_feed_upload",
    )
    such_upload = st.file_uploader(
        "Suchbegriffe-Datei",
        type=["csv"],
        accept_multiple_files=False,
        key="brand_search_upload",
    )
    ausschluss_upload = st.file_uploader(
        "Ausschlussliste",
        type=["csv"],
        accept_multiple_files=False,
        key="brand_exclusion_upload",
    )

    st.caption(
        "Ausgewaehlt: "
        f"Feed={'ja' if feed_upload else 'nein'}, "
        f"Suchbegriffe={'ja' if such_upload else 'nein'}, "
        f"Ausschlussliste={'ja' if ausschluss_upload else 'nein'}"
    )

    if "brand_report" not in st.session_state:
        st.session_state["brand_report"] = None

    if st.button("Analyse erstellen", type="primary"):
        if not feed_upload or not such_upload or not ausschluss_upload:
            st.error("Bitte alle drei Dateien hochladen (Feed, Suchbegriffe, Ausschlussliste).")
        else:
            progress = st.progress(0.0)
            with st.spinner("Brand-Auswertung wird erstellt..."):
                try:
                    brand_counts = lese_brand_counts(feed_upload.getvalue(), feed_upload.name)
                    progress.progress(0.2)

                    def _progress(done: int, total: int) -> None:
                        if total > 0:
                            progress.progress(min(0.2 + (done / total) * 0.6, 0.8))

                    stats, timeframe = aggregiere_suchbegriffe(
                        such_upload.getvalue(),
                        such_upload.name,
                        brand_counts,
                        progress_cb=_progress,
                    )
                    progress.progress(0.85)
                    exclude_set = lese_ausschlussliste(
                        ausschluss_upload.getvalue(),
                        ausschluss_upload.name,
                    )

                    ergebnisse = []
                    for brand, data in stats.items():
                        produkte = brand_counts.get(brand, data["produkte"])
                        klicks = data["klicks"]
                        kosten = data["kosten_eur"]
                        convs = data["conversions"]
                        conv_wert = data["conv_wert_eur"]
                        cpc = kosten / klicks if klicks > 0 else 0.0
                        roas = conv_wert / kosten if kosten > 0 else 0.0
                        ergebnisse.append(
                            (brand, produkte, klicks, cpc, kosten, convs, conv_wert, roas)
                        )
                    ergebnisse.sort(key=lambda value: value[4], reverse=True)

                    report_bytes = schreibe_excel_bytes(
                        ergebnisse,
                        stats,
                        brand_counts,
                        exclude_set,
                        timeframe,
                    )
                    report_name = _resolve_output_file_name(output_file_name)
                    st.session_state["brand_report"] = {
                        "report_bytes": report_bytes,
                        "report_name": report_name,
                    }
                    progress.progress(1.0)
                    st.success("Analyse erfolgreich erstellt.")
                except Exception as exc:  # noqa: BLE001
                    st.session_state["brand_report"] = None
                    st.error(f"Fehler beim Erstellen der Brand-Auswertung: {exc}")

    result = st.session_state.get("brand_report")
    if result:
        st.download_button(
            label="Excel-Report herunterladen",
            data=result["report_bytes"],
            file_name=result["report_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


_render_ui()
